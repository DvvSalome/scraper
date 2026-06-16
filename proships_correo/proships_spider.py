# -*- coding: utf-8 -*-
"""proships_spider_plantilla_xlsx.py

✅ Cambios solicitados
- SOLO Proships (se elimina PackTrack/Paquetes).
- En vez de resultados.csv, escribe en el archivo Excel plantilla:
    FORMATO DE NOVEDADES SAC.xlsx
  conservando estilos (solo se setean valores).

📌 Columnas en la plantilla (encabezados):
- GUIA  PROSHIPS   <= tracking
- NOMBRE           <= buyer
- DIRECCION        <= address
- TELEFONO         <= phone
- CIUDAD Y DEP.    <= city (extraída de un TH cuyo texto sea "city")

🔐 Credenciales
- PROSHIPS_EMAIL / PROSHIPS_PASSWORD en variables de entorno.
"""

import os
import time
from copy import copy
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import chromedriver_autoinstaller
# ===================== CONFIG =====================
LOGIN_URL_PROSHIPS = "https://tracking.proships.com/login"
SHIPMENTS_URL = "https://tracking.proships.com/shipments"

EMAIL_PRO = os.getenv("PROSHIPS_EMAIL", "jara.171201@gmail.com")
PASSWORD_PRO = os.getenv("PROSHIPS_PASSWORD", "Ontime01")

SEARCH_INPUT = (
    By.CSS_SELECTOR,
    "#top-search, input#search, input[placeholder*='Search'], input[placeholder*='Buscar']",
)
TIMELINE_CONTAINER = (By.CSS_SELECTOR, ".timeline-items, .shipment-timeline, .timeline")
TABLE_SELECTOR = (By.CSS_SELECTOR, "table.table, table.table-hover, table.table-striped, table.table-condensed")

PAQUETES_FILE = "paquetes.txt"  # lista de guías a consultar (1 por línea)

# Plantilla Excel (se conserva estilo)
PLANTILLA_XLSX = "FORMATO DE NOVEDADES SAC.xlsx"
PLANTILLA_SHEET = os.getenv("PLANTILLA_SHEET", "")  # opcional: si no está, usa la activa

HEADLESS = False
# ===================================================


# ---------------- Driver helpers -------------------
def build_driver(headless: bool = False):
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--lang=es-ES,es;q=0.9")
    options.add_argument("--window-size=1366,900")
    options.add_argument("--disable-blink-features=AutomationControlled")

    if headless:
        options.add_argument("--headless=new")

    service = Service(chromedriver_autoinstaller.install())

    driver = webdriver.Chrome(
        service=service,
        options=options
    )

    driver.set_page_load_timeout(90)
    driver.set_script_timeout(60)
    return driver


def wait_for(drv, timeout=30):
    return WebDriverWait(drv, timeout)


def robust_get(driver, url: str, wait_selector=None, timeout: int = 45, label: str = ""):
    print(f"➡️  Navegando a {label or url}")
    try:
        driver.get(url)
    except TimeoutException:
        try:
            ready = driver.execute_script("return document.readyState")
        except WebDriverException:
            ready = "unknown"
        if ready not in ("interactive", "complete"):
            ts = int(time.time())
            try:
                driver.save_screenshot(f"/tmp/nav_timeout_{ts}.png")
            except Exception:
                pass
            try:
                with open(f"/tmp/nav_timeout_{ts}.html", "w", encoding="utf-8") as f:
                    f.write(driver.page_source or "")
            except Exception:
                pass
            time.sleep(2)
            driver.get(url)
    if wait_selector:
        wait_for(driver, timeout).until(EC.presence_of_element_located(wait_selector))


# ---------------- Proships ------------------------
def login_proships(driver):
    if not EMAIL_PRO or not PASSWORD_PRO:
        # Mantengo defaults vacíos para que no se quemen credenciales en código.
        # Si en tu ambiente ya usas defaults, puedes volver a ponerlos.
        print("⚠️  PROSHIPS_EMAIL/PROSHIPS_PASSWORD no están en variables de entorno.")

    robust_get(
        driver,
        LOGIN_URL_PROSHIPS,
        wait_selector=(By.CSS_SELECTOR, "input[name='email'], input[type='email']"),
        label="Login Proships",
    )
    w = wait_for(driver, 40)

    email_el = w.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[name='email'], input[type='email']")))
    email_el.clear()
    email_el.send_keys(EMAIL_PRO)

    pass_el = w.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[name='password'], input[type='password']")))
    pass_el.clear()
    pass_el.send_keys(PASSWORD_PRO)

    try:
        driver.find_element(By.CSS_SELECTOR, 'button[type="submit"], input[type="submit"]').click()
    except NoSuchElementException:
        pass_el.send_keys(Keys.ENTER)

    w.until(lambda d: "/login" not in d.current_url)
    print("✅ Login Proships OK")


def _find_search_input(driver):
    inputs = driver.find_elements(*SEARCH_INPUT)
    for el in inputs:
        if el.is_displayed() and el.is_enabled():
            return el
    for sel in ["input[name='search']", "input[name='q']"]:
        found = driver.find_elements(By.CSS_SELECTOR, sel)
        for el in found:
            if el.is_displayed() and el.is_enabled():
                return el
    raise NoSuchElementException("No se encontró el campo de búsqueda en Proships.")


def proships_buscar(driver, tracking: str):
    try:
        inp = _find_search_input(driver)
    except Exception:
        robust_get(driver, SHIPMENTS_URL, wait_selector=SEARCH_INPUT, timeout=45, label="Shipments Proships")
        inp = _find_search_input(driver)

    inp.clear()
    inp.send_keys(tracking)
    inp.send_keys(Keys.ENTER)

    # Espera timeline o resultados
    try:
        wait_for(driver, 25).until(EC.presence_of_element_located(TIMELINE_CONTAINER))
    except TimeoutException:
        time.sleep(1.0)


def proships_extraer_timeline(driver) -> Dict[str, str]:
    data = {"checkpoint_code": ""}
    try:
        first_item = driver.find_element(
            By.CSS_SELECTOR,
            ".timeline-items .row .timeline-item, .timeline-items .timeline-item, .timeline .timeline-item",
        )
    except Exception:
        return data

    for sel in ["address.checkpoint-code", ".checkpoint-code", ".checkpoint, address"]:
        try:
            chk = first_item.find_element(By.CSS_SELECTOR, sel)
            text = (chk.text or "").strip()
            if text:
                data["checkpoint_code"] = text
                break
        except Exception:
            continue
    return data



def proships_extraer_contacto(driver) -> Dict[str, str]:
    """Extrae buyer/phone/address desde la página.

    ✅ Primero intenta por tablas con <th>/<td> (lo más común en Proships).
    ✅ Si no aparece, hace fallback buscando labels en cualquier parte del DOM.
    """
    out = {"buyer": "", "phone": "", "address": ""}

    def try_th_td():
        # Busca cualquier TH que contenga el label y toma el TD asociado.
        def pick(label_variants):
            ths = driver.find_elements(By.CSS_SELECTOR, "th")
            for th in ths:
                txt = (th.text or "").strip().lower()
                if any(v in txt for v in label_variants):
                    # intenta mismo TR
                    try:
                        tr = th.find_element(By.XPATH, "./ancestor::tr[1]")
                        tds = tr.find_elements(By.CSS_SELECTOR, "td")
                        if tds:
                            val = (tds[0].text or "").strip()
                            if val:
                                return val
                    except Exception:
                        pass
                    # siguiente td
                    try:
                        nxt = th.find_element(By.XPATH, "following::td[1]")
                        val = (nxt.text or "").strip()
                        if val:
                            return val
                    except Exception:
                        pass
            return ""

        out["buyer"] = out["buyer"] or pick(["buyer", "comprador", "cliente", "name"])
        out["phone"] = out["phone"] or pick(["phone", "tel", "teléfono", "telefono", "celular", "mobile"])
        out["address"] = out["address"] or pick(["address", "dirección", "direccion", "addr"])

    def try_generic_labels():
        # Fallback: busca elementos que parezcan label y toma su siguiente hermano con texto.
        pairs = driver.find_elements(
            By.CSS_SELECTOR,
            "[class*='label'], [class*='field'], [class*='row'], [class*='col']",
        )

        def grab_after(el):
            # intenta siguiente sibling
            try:
                sib = el.find_element(By.XPATH, "following-sibling::*[1]")
                val = (sib.text or "").strip()
                return val
            except Exception:
                return ""

        for el in pairs:
            try:
                lbl = (el.text or "").strip().lower()
            except Exception:
                continue
            if not lbl or len(lbl) > 40:
                continue

            if (not out["buyer"]) and ("buyer" in lbl or "comprador" in lbl or "cliente" == lbl):
                val = grab_after(el)
                if val:
                    out["buyer"] = val

            if (not out["phone"]) and ("phone" in lbl or "tel" in lbl or "celular" in lbl):
                val = grab_after(el)
                if val:
                    out["phone"] = val

            if (not out["address"]) and ("address" in lbl or "dirección" in lbl or "direccion" in lbl):
                val = grab_after(el)
                if val:
                    out["address"] = val

            if out["buyer"] and out["phone"] and out["address"]:
                break

    # Espera algo de contenido
    try:
        wait_for(driver, 10).until(lambda d: (d.page_source or "").strip() != "")
    except Exception:
        pass

    try_th_td()
    if not (out["buyer"] and out["phone"] and out["address"]):
        try_generic_labels()

    return out
    etiquetas = ["buyer", "phone", "tel", "teléfono", "address", "dirección"]
    tablas = driver.find_elements(By.CSS_SELECTOR, "table")

    tabla_objetivo = None
    for tbl in tablas:
        try:
            txt = (tbl.get_attribute("innerText") or "").lower()
        except Exception:
            continue
        if any(lbl in txt for lbl in etiquetas):
            tabla_objetivo = tbl
            break

    if tabla_objetivo is None:
        return out

    filas = tabla_objetivo.find_elements(By.CSS_SELECTOR, "tr")
    for tr in filas:
        celdas = tr.find_elements(By.CSS_SELECTOR, "th,td")
        if len(celdas) < 2:
            continue
        label = (celdas[0].text or "").strip().lower()
        valor_td = (celdas[1].text or "").strip()

        if not out["buyer"] and "buyer" in label:
            out["buyer"] = valor_td
        elif not out["phone"] and ("phone" in label or "tel" in label or "teléfono" in label):
            out["phone"] = valor_td
        elif not out["address"] and ("address" in label or "dirección" in label):
            out["address"] = valor_td

    return out


def proships_extraer_city(driver) -> str:
    """Busca un TH con texto 'city' y retorna su TD asociado."""
    # Intento 1: con DOM (tablas)
    try:
        ths = driver.find_elements(By.CSS_SELECTOR, "th")
        for th in ths:
            try:
                if (th.text or "").strip().lower() == "city":
                    tr = None
                    try:
                        tr = th.find_element(By.XPATH, "./ancestor::tr[1]")
                    except Exception:
                        tr = None

                    if tr is not None:
                        tds = tr.find_elements(By.CSS_SELECTOR, "td")
                        if tds:
                            return (tds[0].text or "").strip()

                    # si no hay TR claro, toma el siguiente TD
                    try:
                        nxt = th.find_element(By.XPATH, "following::td[1]")
                        return (nxt.text or "").strip()
                    except Exception:
                        pass
            except Exception:
                continue
    except Exception:
        pass

    # Intento 2: fallback con page_source (por si Proships renderiza raro)
    html = driver.page_source or ""
    low = html.lower()
    if "<th" in low and "city" in low:
        # fallback básico sin bs4 para no agregar dependencia
        # Busca el primer 'city' y luego el siguiente <td>...</td>
        idx = low.find(">city<")
        if idx != -1:
            td_start = low.find("<td", idx)
            if td_start != -1:
                td_gt = low.find(">", td_start)
                td_end = low.find("</td>", td_gt)
                if td_gt != -1 and td_end != -1:
                    return html[td_gt + 1 : td_end].strip()

    return ""


def proships_procesar(driver, tracking: str) -> Dict[str, str]:
    proships_buscar(driver, tracking)
    timeline = proships_extraer_timeline(driver)
    contacto = proships_extraer_contacto(driver)
    city = proships_extraer_city(driver)

    return {
        "tracking": tracking,
        "checkpoint_code": timeline.get("checkpoint_code", ""),
        "buyer": contacto.get("buyer", ""),
        "phone": contacto.get("phone", ""),
        "address": contacto.get("address", ""),
        "city": city,
    }


# ---------------- Excel (plantilla con estilo) -----------------
def leer_paquetes() -> List[str]:
    if Path(PAQUETES_FILE).exists():
        with open(PAQUETES_FILE, "r", encoding="utf-8") as f:
            return [line.strip() for line in f if line.strip()]
    raise RuntimeError(f"No se encontró el archivo {PAQUETES_FILE}")


def _normalizar_header(s: str) -> str:
    return (s or "").strip().lower()


def _mapear_columnas_por_header(ws, max_scan_rows: int = 50) -> (Dict[str, int], int):
    """Encuentra la fila de encabezados y retorna:
    - mapping header_normalizado -> columna (1-index)
    - header_row (1-index)

    La plantilla puede tener logos/espacios arriba (por ejemplo encabezados en fila 6).
    Escanea las primeras `max_scan_rows` filas buscando los headers requeridos.
    """
    required = {
        "guia  proships",
        "guia proships",
        "nombre",
        "direccion",
        "teléfono",
        "telefono",
        "ciudad y dep.",
        "ciudad y dep",
        "novedad operador",
    }


    def is_required(key: str) -> bool:
        k = key.replace("  ", " ").strip().lower()
        return key in required or k in required

    best_row = None
    best_mapping: Dict[str, int] = {}

    scan_rows = min(ws.max_row, max_scan_rows)
    for row in range(1, scan_rows + 1):
        mapping: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            key = _normalizar_header(str(v))
            if not key:
                continue
            if is_required(key):
                # normalizamos claves canónicas para uso interno
                canon = key.replace("  ", " ").strip().lower()
                mapping[canon] = col

        # contamos cuántos de los esenciales encontramos
        essentials = {
            "guia proships",
            "nombre",
            "direccion",
            "telefono",
            "ciudad y dep.",
            "novedad operador",
        }
        found = set()
        for k in mapping.keys():
            if k.startswith("guia proships"):
                found.add("guia proships")
            elif k == "nombre":
                found.add("nombre")
            elif k == "direccion":
                found.add("direccion")
            elif k in ("telefono", "teléfono"):
                found.add("telefono")
            elif k in ("ciudad y dep.", "ciudad y dep"):
                found.add("ciudad y dep.")

        if len(found) > len(best_mapping):
            best_mapping = mapping
            best_row = row

        if len(found) == 5:
            return mapping, row

    if best_row is None or not best_mapping:
        raise RuntimeError(
            "No encontré la fila de encabezados en el Excel (busqué en las primeras "
            f"{scan_rows} filas). Asegúrate que existan: GUIA PROSHIPS, NOMBRE, DIRECCION, TELEFONO, CIUDAD Y DEP."
        )

    return best_mapping, best_row



def _encontrar_primera_fila_vacia(ws, col_idx: int, start_row: int = 2) -> int:
    """Busca primera fila vacía en una columna (por defecto debajo del header)."""
    r = start_row
    while True:
        cell = _celda_editable(ws, r, col_idx)
        if cell.value in (None, ""):
            return r
        r += 1


def _celda_editable(ws, row: int, col: int):
    """Devuelve una celda editable. Si (row,col) cae en un merge, retorna la esquina sup-izq del rango."""
    c = ws.cell(row=row, column=col)
    # openpyxl usa MergedCell (read-only) para celdas no-top-left dentro de un merge.
    if c.__class__.__name__ == "MergedCell":
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return ws.cell(row=rng.min_row, column=rng.min_col)
    return c


def _copiar_estilo_fila(ws, src_row: int, dst_row: int, col_indices: List[int]):
    """Copia estilos/celda desde src_row a dst_row para las columnas dadas."""
    for col in col_indices:
        src = _celda_editable(ws, src_row, col)
        dst = _celda_editable(ws, dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)



def limpiar_datos_plantilla_xlsx():
    """Limpia (solo valores) todas las filas de datos debajo del header en FORMATO DE NOVEDADES SAC.xlsx.
    No toca estilos, bordes ni merges (usa _celda_editable).
    """
    if not Path(PLANTILLA_XLSX).exists():
        raise RuntimeError(f"No se encontró la plantilla: {PLANTILLA_XLSX}")

    wb = load_workbook(PLANTILLA_XLSX)
    ws = wb[PLANTILLA_SHEET] if PLANTILLA_SHEET and PLANTILLA_SHEET in wb.sheetnames else wb.active

    headers, header_row = _mapear_columnas_por_header(ws)

    col_guia = headers.get("guia proships")
    col_nombre = headers.get("nombre")
    col_dir = headers.get("direccion")
    col_tel = headers.get("telefono") or headers.get("teléfono")
    col_city = headers.get("ciudad y dep.") or headers.get("ciudad y dep")

    cols = [c for c in [col_guia, col_nombre, col_dir, col_tel, col_city] if c]

    start_row = header_row + 1
    # Limpia hasta la última fila usada por la hoja (max_row)
    for r in range(start_row, ws.max_row + 1):
        # Si la fila está totalmente vacía en estas columnas, igual la limpiamos (no rompe nada)
        for c in cols:
            cell = _celda_editable(ws, r, c)
            cell.value = ""

    wb.save(PLANTILLA_XLSX)



def escribir_en_plantilla_xlsx(rowdata: Dict[str, str]):
    """Escribe en la plantilla Excel conservando estilo."""
    if not Path(PLANTILLA_XLSX).exists():
        raise RuntimeError(f"No se encontró la plantilla: {PLANTILLA_XLSX}")

    wb = load_workbook(PLANTILLA_XLSX)
    ws = wb[PLANTILLA_SHEET] if PLANTILLA_SHEET and PLANTILLA_SHEET in wb.sheetnames else wb.active

    headers, header_row = _mapear_columnas_por_header(ws)

    # Columnas esperadas (canónicas)
    col_guia = headers.get("guia proships")
    col_nombre = headers.get("nombre")
    col_dir = headers.get("direccion")
    col_tel = headers.get("telefono") or headers.get("teléfono")
    col_city = headers.get("ciudad y dep.") or headers.get("ciudad y dep")
    col_novedad = headers.get("novedad operador")


    missing = []
    if not col_guia: missing.append("GUIA PROSHIPS")
    if not col_nombre: missing.append("NOMBRE")
    if not col_dir: missing.append("DIRECCION")
    if not col_tel: missing.append("TELEFONO")
    if not col_city: missing.append("CIUDAD Y DEP.")
    if not col_novedad:
        missing.append("NOVEDAD OPERADOR")

    if missing:
        raise RuntimeError(
            "No encontré estos encabezados en el Excel: "
            + ", ".join(missing)
            + f"\nEncontré fila de encabezados en la fila {header_row}, pero faltan columnas."
        )

    # Primera fila vacía según la columna GUIA, empezando debajo de los headers
    next_row = _encontrar_primera_fila_vacia(ws, col_guia, start_row=header_row + 1)

    # Copiar estilo desde la fila anterior si existe (o desde fila 2)
    style_src_row = next_row - 1 if next_row > 2 else 2
    _copiar_estilo_fila(
        ws,
        style_src_row,
        next_row,
        [col_guia, col_nombre, col_dir, col_tel, col_city, col_novedad]
    )

    _celda_editable(ws, next_row, col_guia).value = (rowdata.get("tracking") or "").strip()
    _celda_editable(ws, next_row, col_nombre).value = (rowdata.get("buyer") or "").strip()
    _celda_editable(ws, next_row, col_dir).value = (rowdata.get("address") or "").strip()
    _celda_editable(ws, next_row, col_tel).value = (rowdata.get("phone") or "").strip()
    _celda_editable(ws, next_row, col_city).value = (rowdata.get("city") or "").strip()
    _celda_editable(ws, next_row, col_novedad).value = (
        rowdata.get("checkpoint_code") or ""
    ).strip()


    # Forzar fondo blanco en la fila de datos (sin tocar bordes/fuentes)
    blanco = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    for cidx in [col_guia, col_nombre, col_dir, col_tel, col_city, col_novedad]:
        cell = _celda_editable(ws, next_row, cidx)
        cell.fill = copy(blanco)


    wb.save(PLANTILLA_XLSX)


# ---------------- Main flow -----------------------
def main():
    trackings = leer_paquetes()
    # 🔄 Limpia el FORMATO DE NOVEDADES antes de escribir (no acumula datos)
    limpiar_datos_plantilla_xlsx()
    driver = build_driver(headless=HEADLESS)

    try:
        login_proships(driver)

        total = len(trackings)
        for i, trk in enumerate(trackings, start=1):
            print(f"\n=========== [Proships {i}/{total}] {trk} ===========")

            pro = proships_procesar(driver, trk)
            escribir_en_plantilla_xlsx(pro)

            # pausa leve para no cargar demasiado
            time.sleep(1.0)

    finally:
        try:
            driver.quit()
        except Exception:
            pass

    print(f"\n📘 Listo: datos escritos en {PLANTILLA_XLSX}")


if __name__ == "__main__":
    main()
